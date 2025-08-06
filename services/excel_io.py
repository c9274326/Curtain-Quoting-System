# services/excel_io.py

import os
import zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime

class ExcelManager:
    """Excel 檔案管理器"""

    def __init__(self, config):
        # 接收並保存設定，以便後續擴充使用
        self.config = config

    def create_price_template(self, filename: str):
        """建立一份合法的 .xlsx 價格表範本"""
        dir_name = os.path.dirname(filename)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "價格表"
        headers = ['窗簾類型', '材質', '單價', '單位', '最低數量']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 範例資料
        sample = [
            ['蛇行布簾', '國產遮光布', 450, '尺', 8],
            ['捲簾',     '遮光布',     250, '才', 15]
        ]
        for r, row in enumerate(sample, start=2):
            for c, v in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=v)

        wb.save(filename)

    def create_quote_from_template(self, template_path: str, output_path: str, quote_data: dict):
        """
        以客戶專屬範本產生報價單；若範本不存在或非合法 .xlsx，
        自動用內建範本取代再產生報價。
        """
        # 1. 範本檢查：存在、.xlsx 副檔名、且為 zip archive
        use_path = template_path
        if (not os.path.exists(template_path)
           or not template_path.lower().endswith('.xlsx')
           or not zipfile.is_zipfile(template_path)):
            # 產生內建範本
            default_tpl = os.path.join(os.path.dirname(output_path), 'default_quote_template.xlsx')
            self.create_price_template(default_tpl)
            use_path = default_tpl

        # 2. 嘗試載入範本
        try:
            wb = load_workbook(use_path)
        except InvalidFileException:
            wb = Workbook()
        ws = wb.active

        # 3. 寫入本公司資訊
        ws['B2'] = quote_data['company']['company_name']
        ws['B3'] = quote_data['company']['company_phone']
        ws['B4'] = quote_data['company']['company_address']

        # 4. 寫入客戶資訊
        ws['D2'] = quote_data['customer']['name']
        ws['D3'] = quote_data['customer']['phone']
        ws['D4'] = quote_data['customer']['address']

        # 5. 寫入明細（起始列 7），並正確帶入 unit
        start_row = 7
        for idx, item in enumerate(quote_data['items'], start=start_row):
            ws.cell(row=idx, column=1, value=item['item'])
            ws.cell(row=idx, column=2, value=item['quantity'])
            ws.cell(row=idx, column=3, value=item['unit'])
            ws.cell(row=idx, column=4, value=item['unit_price'])
            ws.cell(row=idx, column=5, value=item['subtotal'])
            # 如果有日期欄（第6欄），也可：
            if 'date' in item:
                ws.cell(row=idx, column=6, value=item['date'])

        # 6. 寫入總計（假設在 E20–E22）
        ws['E20'] = quote_data['subtotal']
        ws['E21'] = quote_data['total']

        # 7. 確保輸出目錄存在
        out_dir = os.path.dirname(output_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        wb.save(output_path)
